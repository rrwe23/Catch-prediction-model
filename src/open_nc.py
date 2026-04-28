import xarray as xr
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── 제주 갈치 어장 범위 ──────────────────────────
LAT_MIN, LAT_MAX = 32.0, 34.0
LON_MIN, LON_MAX = 125.0, 127.5

# ── 파일 선택 ────────────────────────────────────
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="NC 파일 선택",
    filetypes=[("NetCDF files", "*.nc"), ("All files", "*.*")]
)

if not file_path:
    print("파일을 선택하지 않았습니다.")
    exit()

print(f"파일 로드 중: {os.path.basename(file_path)}")

# ── NC 파일 로드 ─────────────────────────────────
ds = xr.open_dataset(file_path)
print("\n[전체 데이터셋 구조]")
print(ds)

# 위경도 좌표명 자동 감지
lat_names = ["latitude", "lat", "y", "nav_lat"]
lon_names = ["longitude", "lon", "x", "nav_lon"]

lat_dim = next((n for n in lat_names if n in ds.coords or n in ds.dims), None)
lon_dim = next((n for n in lon_names if n in ds.coords or n in ds.dims), None)

if not lat_dim or not lon_dim:
    print(f"\n좌표 목록: {list(ds.coords)}")
    lat_dim = input("위도 좌표명 입력: ").strip()
    lon_dim = input("경도 좌표명 입력: ").strip()

print(f"\n위도 좌표: {lat_dim}, 경도 좌표: {lon_dim}")

# ── 범위 필터링 ──────────────────────────────────
ds_filtered = ds.sel(
    {lat_dim: slice(LAT_MIN, LAT_MAX),
     lon_dim: slice(LON_MIN, LON_MAX)}
)

if ds_filtered[lat_dim].size == 0 or ds_filtered[lon_dim].size == 0:
    print("slice 방식 실패 → where 방식으로 재시도")
    lat_mask = (ds[lat_dim] >= LAT_MIN) & (ds[lat_dim] <= LAT_MAX)
    lon_mask = (ds[lon_dim] >= LON_MIN) & (ds[lon_dim] <= LON_MAX)
    ds_filtered = ds.where(lat_mask & lon_mask, drop=True)

print(f"\n[필터링 결과]")
print(f"  위도 포인트 수: {ds_filtered[lat_dim].size}")
print(f"  경도 포인트 수: {ds_filtered[lon_dim].size}")

# ── 모든 변수 DataFrame 변환 ─────────────────────
variables = list(ds_filtered.data_vars)
print(f"\n[변수 목록] {variables}")

df_list = []

for var in variables:
    da = ds_filtered[var]

    # time, depth 첫 번째 슬라이스
    if "time" in da.dims:
        da = da.isel(time=0)
    if "depth" in da.dims:
        da = da.isel(depth=0)

    df_var = da.to_dataframe().reset_index()
    df_var = df_var[[lat_dim, lon_dim, var]].dropna(subset=[var])
    df_list.append(df_var.set_index([lat_dim, lon_dim]))
    print(f"  {var}: {len(df_var):,}개 포인트")

# 모든 변수 병합 (위경도 기준 join)
df = pd.concat(df_list, axis=1).reset_index()
df = df.sort_values([lat_dim, lon_dim]).reset_index(drop=True)

# 컬럼명 정리
df = df.rename(columns={lat_dim: "위도(latitude)", lon_dim: "경도(longitude)"})

print(f"\n총 {len(df):,}개 좌표 포인트 / {len(variables)}개 변수")

# ── 엑셀 저장 ────────────────────────────────────
save_path = file_path.replace(".nc", "_jeju_zone.xlsx")
wb = Workbook()

COLOR_HEADER_BG   = "1A3A5C"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_INFO_BG     = "E8F4F8"
COLOR_INFO_FONT   = "0D2444"
COLOR_BORDER      = "BDD7EE"

thin   = Side(style="thin", color=COLOR_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 시트 1: 전체 데이터 ──────────────────────────
ws_data = wb.active
ws_data.title = "갈치어장_좌표데이터"

total_cols = len(df.columns)

ws_data.merge_cells(f"A1:{get_column_letter(total_cols)}1")
c = ws_data["A1"]
c.value = "제주 갈치 어장 해역 좌표 데이터 — 전체 변수"
c.font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FONT)
c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_data.row_dimensions[1].height = 28

ws_data.merge_cells(f"A2:{get_column_letter(total_cols)}2")
c = ws_data["A2"]
c.value = (f"분석 범위: 위도 {LAT_MIN}°N ~ {LAT_MAX}°N  /  경도 {LON_MIN}°E ~ {LON_MAX}°E  "
           f"|  총 {len(df):,}개 포인트  |  변수: {', '.join(variables)}")
c.font = Font(name="Arial", size=10, color=COLOR_INFO_FONT)
c.fill = PatternFill("solid", start_color=COLOR_INFO_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_data.row_dimensions[2].height = 20

for col_idx, header in enumerate(df.columns, 1):
    c = ws_data.cell(row=3, column=col_idx, value=header)
    c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", start_color="2E6B9E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws_data.row_dimensions[3].height = 18

for row_idx, row in enumerate(df.itertuples(index=False), 4):
    for col_idx, value in enumerate(row, 1):
        c = ws_data.cell(row=row_idx, column=col_idx, value=value)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        if row_idx % 2 == 0:
            c.fill = PatternFill("solid", start_color="F0F7FC")

for col_idx in range(1, total_cols + 1):
    ws_data.column_dimensions[get_column_letter(col_idx)].width = 16

ws_data.freeze_panes = "A4"

# ── 시트 2: 변수별 통계 요약 ─────────────────────
ws_stat = wb.create_sheet("변수별_통계")

stat_headers = ["변수", "최솟값", "최댓값", "평균값", "표준편차", "유효 포인트 수"]
for col_idx, h in enumerate(stat_headers, 1):
    c = ws_stat.cell(row=1, column=col_idx, value=h)
    c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center")
    c.border = border
    ws_stat.column_dimensions[get_column_letter(col_idx)].width = 18

for row_idx, var in enumerate(variables, 2):
    series = df[var].dropna()
    stats = [
        var,
        f"{series.min():.4f}",
        f"{series.max():.4f}",
        f"{series.mean():.4f}",
        f"{series.std():.4f}",
        f"{len(series):,}",
    ]
    for col_idx, val in enumerate(stats, 1):
        c = ws_stat.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        if row_idx % 2 == 0:
            c.fill = PatternFill("solid", start_color="F0F7FC")

# ── 시트 3: 메타 정보 ────────────────────────────
ws_meta = wb.create_sheet("메타정보")
ws_meta.column_dimensions["A"].width = 28
ws_meta.column_dimensions["B"].width = 44

meta_rows = [
    ("항목", "값"),
    ("NC 파일명", os.path.basename(file_path)),
    ("전체 변수 목록", ", ".join(variables)),
    ("위도 최소 (°N)", LAT_MIN),
    ("위도 최대 (°N)", LAT_MAX),
    ("경도 최소 (°E)", LON_MIN),
    ("경도 최대 (°E)", LON_MAX),
    ("총 좌표 포인트 수", len(df)),
    ("실제 위도 범위", f"{df['위도(latitude)'].min():.4f} ~ {df['위도(latitude)'].max():.4f}"),
    ("실제 경도 범위", f"{df['경도(longitude)'].min():.4f} ~ {df['경도(longitude)'].max():.4f}"),
    ("목적", "제주 갈치 어장 CNN 예측 모델 입력 데이터"),
    ("어장 기준", "갈치 주어장 (KOSIS 어업생산동향조사 기준)"),
]

for row_idx, (key, val) in enumerate(meta_rows, 1):
    ca = ws_meta.cell(row=row_idx, column=1, value=key)
    cb = ws_meta.cell(row=row_idx, column=2, value=val)
    if row_idx == 1:
        for c in [ca, cb]:
            c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
            c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
            c.alignment = Alignment(horizontal="center")
    else:
        ca.font = Font(name="Arial", bold=True, size=10, color=COLOR_INFO_FONT)
        ca.fill = PatternFill("solid", start_color=COLOR_INFO_BG)
        cb.font = Font(name="Arial", size=10)
        if row_idx % 2 == 0:
            cb.fill = PatternFill("solid", start_color="F0F7FC")
    for c in [ca, cb]:
        c.border = border
        c.alignment = Alignment(vertical="center", horizontal=c.alignment.horizontal or "left")
    ws_meta.row_dimensions[row_idx].height = 18

wb.save(save_path)
print(f"\n엑셀 저장 완료: {save_path}")
messagebox.showinfo("완료", f"저장 완료!\n\n{save_path}")