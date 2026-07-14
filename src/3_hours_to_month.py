"""
CMEMS WAVE 파일 처리기

[입력]
  cmems_mod_glo_wav_my_*.nc  (3시간 간격 데이터, ~58,440 timestamps)

[처리]
  1. 3시간 간격 데이터를 월별로 집계
  2. 각 격자 포인트의 월평균 파고 계산
  3. CSV로 저장 (기존 형식 유지)

[출력]
  cmems_wav_jeju_monthly.csv  (월별 격자별 파고)
  cmems_wav_jeju_timeseries.xlsx (엑셀, 시계열 + 월별 공간평균 + 통계 + 메타)
"""

import xarray as xr
import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── 제주 갈치 어장 범위 ──────────────────────────
LAT_MIN, LAT_MAX = 32.0, 34.0
LON_MIN, LON_MAX = 125.0, 127.5

# ── 파일 선택 ────────────────────────────────────
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="WAVE NC 파일 선택 (cmems_mod_glo_wav_*.nc)",
    filetypes=[("NetCDF files", "*.nc"), ("All files", "*.*")]
)

if not file_path:
    print("파일을 선택하지 않았습니다.")
    exit()

print(f"파일 로드 중: {os.path.basename(file_path)}")

# ── NC 파일 로드 ─────────────────────────────────
ds = xr.open_dataset(file_path)
print("\n[원본 데이터 구조]")
print(ds)

# ── 위경도 범위 필터링 ──────────────────────────
ds_filt = ds.sel(
    latitude=slice(LAT_MIN, LAT_MAX),
    longitude=slice(LON_MIN, LON_MAX),
)

print(f"\n[필터링 결과]")
print(f"  위도: {ds_filt['latitude'].size}개  ({float(ds_filt['latitude'].min()):.2f}~{float(ds_filt['latitude'].max()):.2f})")
print(f"  경도: {ds_filt['longitude'].size}개  ({float(ds_filt['longitude'].min()):.2f}~{float(ds_filt['longitude'].max()):.2f})")
print(f"  시간: {ds_filt['time'].size:,}개 (3시간 간격)")

# ── 월별 평균 집계 ⭐ 핵심 ───────────────────────
print(f"\n[월별 평균 집계 중...]")
ds_monthly = ds_filt.resample(time="MS").mean(skipna=True)

print(f"  집계 완료: {ds_filt['time'].size:,} → {ds_monthly['time'].size}개월")
print(f"  기간: {str(ds_monthly['time'].values[0])[:10]} ~ {str(ds_monthly['time'].values[-1])[:10]}")

# ── DataFrame 변환 ───────────────────────────────
variables = list(ds_monthly.data_vars)
print(f"\n[변수 목록] {variables}")

df_list = []
for var in variables:
    da = ds_monthly[var]
    df_var = da.to_dataframe().reset_index()

    # 차원이 (time, lat, lon)인 경우만 처리
    keep = [c for c in ["time", "latitude", "longitude", var] if c in df_var.columns]
    df_var = df_var[keep].dropna(subset=[var])
    df_list.append(df_var.set_index([c for c in keep if c != var]))
    print(f"  {var}: {len(df_var):,}개 레코드")

df = pd.concat(df_list, axis=1).reset_index()

# 시간 포맷
df["time"] = pd.to_datetime(df["time"]).dt.strftime("%Y-%m")

# 컬럼 순서/이름
df = df.rename(columns={
    "time":      "시간(time)",
    "latitude":  "위도(latitude)",
    "longitude": "경도(longitude)",
})
df = df.sort_values(["시간(time)", "위도(latitude)", "경도(longitude)"]).reset_index(drop=True)

print(f"\n총 {len(df):,}행 / {len(variables)}개 변수")

# ── CSV 저장 (간단 버전) ────────────────────────
csv_path = file_path.replace(".nc", "_monthly.csv")
df.to_csv(csv_path, index=False, encoding="utf-8-sig")
print(f"\nCSV 저장: {csv_path}")
print(f"  파일 크기: {os.path.getsize(csv_path) / 1024:.2f} KB")

# ── 엑셀 저장 (기존 형식과 동일) ─────────────────
save_path = file_path.replace(".nc", "_timeseries.xlsx")
wb = Workbook()

COLOR_HEADER_BG   = "1A3A5C"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_INFO_BG     = "E8F4F8"
COLOR_INFO_FONT   = "0D2444"
COLOR_BORDER      = "BDD7EE"

thin   = Side(style="thin", color=COLOR_BORDER)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── 시트 1: 시계열 전체 데이터 ───────────────────
ws_data = wb.active
ws_data.title = "시계열_전체데이터"

total_cols = len(df.columns)
ws_data.merge_cells(f"A1:{get_column_letter(total_cols)}1")
c = ws_data["A1"]
c.value = "제주 갈치 어장 해역 — WAVE 월별 데이터"
c.font = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FONT)
c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_data.row_dimensions[1].height = 28

time_range = f"{df['시간(time)'].iloc[0]} ~ {df['시간(time)'].iloc[-1]}"
ws_data.merge_cells(f"A2:{get_column_letter(total_cols)}2")
c = ws_data["A2"]
c.value = (f"기간: {time_range}  |  위도 {LAT_MIN}°N ~ {LAT_MAX}°N  /  경도 {LON_MIN}°E ~ {LON_MAX}°E  "
           f"|  총 {len(df):,}행  |  변수: {', '.join(variables)}  |  3시간 간격 → 월 평균 집계")
c.font = Font(name="Arial", size=10, color=COLOR_INFO_FONT)
c.fill = PatternFill("solid", start_color=COLOR_INFO_BG)
c.alignment = Alignment(horizontal="left", vertical="center")
ws_data.row_dimensions[2].height = 20

# 헤더
for col_idx, header in enumerate(df.columns, 1):
    c = ws_data.cell(row=3, column=col_idx, value=header)
    c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", start_color="2E6B9E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws_data.row_dimensions[3].height = 18

print("\n엑셀 데이터 쓰는 중...")
for row_idx, row in enumerate(df.itertuples(index=False), 4):
    for col_idx, value in enumerate(row, 1):
        c = ws_data.cell(row=row_idx, column=col_idx, value=value)
        c.font = Font(name="Arial", size=9)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        if row_idx % 2 == 0:
            c.fill = PatternFill("solid", start_color="F0F7FC")

for col_idx in range(1, total_cols + 1):
    ws_data.column_dimensions[get_column_letter(col_idx)].width = 16

ws_data.freeze_panes = "A4"

# ── 시트 2: 월별 공간 평균 ───────────────────────
ws_monthly = wb.create_sheet("월별_공간평균")

df_monthly_avg = df.groupby("시간(time)")[variables].mean().reset_index()
df_monthly_avg.columns = ["시간(time)"] + [f"{v}_평균" for v in variables]

m_cols = len(df_monthly_avg.columns)
ws_monthly.merge_cells(f"A1:{get_column_letter(m_cols)}1")
c = ws_monthly["A1"]
c.value = "월별 공간 평균 (제주 해역 전체 평균)"
c.font = Font(name="Arial", bold=True, size=13, color=COLOR_HEADER_FONT)
c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_monthly.row_dimensions[1].height = 26

for col_idx, header in enumerate(df_monthly_avg.columns, 1):
    c = ws_monthly.cell(row=2, column=col_idx, value=header)
    c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", start_color="2E6B9E")
    c.alignment = Alignment(horizontal="center")
    c.border = border
    ws_monthly.column_dimensions[get_column_letter(col_idx)].width = 18

for row_idx, row in enumerate(df_monthly_avg.itertuples(index=False), 3):
    for col_idx, value in enumerate(row, 1):
        c = ws_monthly.cell(row=row_idx, column=col_idx, value=value)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        if row_idx % 2 == 0:
            c.fill = PatternFill("solid", start_color="F0F7FC")

ws_monthly.freeze_panes = "A3"

# ── 시트 3: 변수별 통계 ──────────────────────────
ws_stat = wb.create_sheet("변수별_통계")

stat_headers = ["변수", "최솟값", "최댓값", "평균값", "표준편차", "유효 레코드 수"]
for col_idx, h in enumerate(stat_headers, 1):
    c = ws_stat.cell(row=1, column=col_idx, value=h)
    c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
    c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center")
    c.border = border
    ws_stat.column_dimensions[get_column_letter(col_idx)].width = 18

for row_idx, var in enumerate(variables, 2):
    series = df[var].dropna()
    stats = [var, f"{series.min():.4f}", f"{series.max():.4f}",
             f"{series.mean():.4f}", f"{series.std():.4f}", f"{len(series):,}"]
    for col_idx, val in enumerate(stats, 1):
        c = ws_stat.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        if row_idx % 2 == 0:
            c.fill = PatternFill("solid", start_color="F0F7FC")

# ── 시트 4: 메타 정보 ────────────────────────────
ws_meta = wb.create_sheet("메타정보")
ws_meta.column_dimensions["A"].width = 28
ws_meta.column_dimensions["B"].width = 44

meta_rows = [
    ("항목", "값"),
    ("NC 파일명", os.path.basename(file_path)),
    ("원본 시간 간격", "3시간"),
    ("집계 방법", "월별 평균 (resample 'MS' + mean)"),
    ("전체 변수 목록", ", ".join(variables)),
    ("기간", time_range),
    ("위도 최소 (°N)", LAT_MIN),
    ("위도 최대 (°N)", LAT_MAX),
    ("경도 최소 (°E)", LON_MIN),
    ("경도 최대 (°E)", LON_MAX),
    ("총 레코드 수 (월별)", len(df)),
    ("실제 위도 범위", f"{df['위도(latitude)'].min():.4f} ~ {df['위도(latitude)'].max():.4f}"),
    ("실제 경도 범위", f"{df['경도(longitude)'].min():.4f} ~ {df['경도(longitude)'].max():.4f}"),
    ("목적", "제주 갈치 어장 CNN 예측 모델 입력 데이터"),
    ("어장 기준", "갈치 주어장 (KOSIS 어업생산동향조사 기준)"),
]

for row_idx, (key, val) in enumerate(meta_rows, 1):
    ca = ws_meta.cell(row=row_idx, column=1, value=key)
    cb = ws_meta.cell(row=row_idx, column=2, value=val)
    if row_idx == 1:
        for c in [ca, cb]:
            c.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
            c.fill = PatternFill("solid", start_color=COLOR_HEADER_BG)
            c.alignment = Alignment(horizontal="center")
    else:
        ca.font = Font(name="Arial", bold=True, size=10, color=COLOR_INFO_FONT)
        ca.fill = PatternFill("solid", start_color=COLOR_INFO_BG)
        cb.font = Font(name="Arial", size=10)
        if row_idx % 2 == 0:
            cb.fill = PatternFill("solid", start_color="F0F7FC")
    for c in [ca, cb]:
        c.border = border
        c.alignment = Alignment(vertical="center", horizontal=c.alignment.horizontal or "left")
    ws_meta.row_dimensions[row_idx].height = 18

wb.save(save_path)
print(f"엑셀 저장: {save_path}")

# ── 월별 집계된 NC 파일도 저장 (preprocess.py에서 활용) ──
nc_out_path = file_path.replace(".nc", "_monthly.nc")
ds_monthly.to_netcdf(nc_out_path)
print(f"월별 NC 저장: {nc_out_path}")

print(f"\n{'='*60}")
print(f"처리 완료!")
print(f"{'='*60}")
print(f"  원본 (3시간 간격): {ds_filt['time'].size:,} 시점")
print(f"  집계 후 (월별):    {len(df_monthly_avg)} 개월")
print(f"  CSV 파일 크기: {os.path.getsize(csv_path) / 1024:.2f} KB")
print(f"\n[저장된 파일]")
print(f"  1. CSV (월별 격자별): {os.path.basename(csv_path)}")
print(f"  2. 엑셀:              {os.path.basename(save_path)}")
print(f"  3. NC (월별 집계):    {os.path.basename(nc_out_path)} ⭐")
print(f"\n→ preprocess.py 에서 #3 NC 파일 사용하면 됩니다")

messagebox.showinfo(
    "완료",
    f"WAVE 데이터 월별 집계 완료!\n\n"
    f"원본: {ds_filt['time'].size:,} 시점 (3시간)\n"
    f"집계: {len(df_monthly_avg)} 개월\n\n"
    f"저장 위치:\n{os.path.dirname(file_path)}"
)